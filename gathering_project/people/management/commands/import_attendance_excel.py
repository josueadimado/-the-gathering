"""
Import people from the \"The Gathering Attendance.xlsx\" file.

Usage:
    python manage.py import_attendance_excel
    python manage.py import_attendance_excel --file \"path/to/file.xlsx\"
"""

from pathlib import Path
import re

from django.core.management.base import BaseCommand, CommandError

from people.models import Person


class Command(BaseCommand):
    help = "Import people from an attendance Excel file into the Person model."

    def add_arguments(self, parser):
        parser.add_argument(
            "--file",
            type=str,
            default="The Gathering Attendance.xlsx",
            help=(
                "Path to the Excel file "
                '(default: "The Gathering Attendance.xlsx" in project root).'
            ),
        )

    def handle(self, *args, **options):
        try:
            import openpyxl  # Imported here so the command fails nicely if missing
        except ImportError as exc:
            raise CommandError(
                "openpyxl is required for this command. "
                "Install it with: pip install openpyxl"
            ) from exc

        file_path = Path(options["file"])
        if not file_path.exists():
            raise CommandError(f"File not found: {file_path}")

        self.stdout.write(self.style.NOTICE(f"Reading Excel file: {file_path}"))

        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active

        # Read headers from first row
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = [str(h).strip() if h is not None else "" for h in header_row]
        header_map = {h.lower(): idx for idx, h in enumerate(headers)}

        # We expect at least Name and Contact columns
        def get_index(*candidates):
            for cand in candidates:
                idx = header_map.get(cand.lower())
                if idx is not None:
                    return idx
            return None

        name_idx = get_index("name")
        country_idx = get_index("country", "country ")
        contact_idx = get_index("contact")

        if name_idx is None or contact_idx is None:
            raise CommandError(
                "Expected columns 'Name' and 'Contact' in the header row. "
                f"Found headers: {headers}"
            )

        created_count = 0
        updated_count = 0
        skipped_count = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            name = (row[name_idx] or "").strip() if row[name_idx] else ""
            country = (
                (row[country_idx] or "").strip()
                if country_idx is not None and row[country_idx]
                else ""
            )
            contact_raw = row[contact_idx]

            if not name and not contact_raw:
                # Completely empty row
                continue

            if not contact_raw:
                skipped_count += 1
                self.stdout.write(
                    self.style.WARNING(f"Skipping row without contact: {row}")
                )
                continue

            # Normalise contact to digits only
            contact_str = re.sub(r"\\D", "", str(contact_raw))
            if not contact_str:
                skipped_count += 1
                self.stdout.write(
                    self.style.WARNING(f"Skipping row with invalid contact: {row}")
                )
                continue

            phone_number = self._to_e164(contact_str, country)

            first_name, last_name = self._split_name(name)

            person, created = Person.objects.get_or_create(
                phone_number=phone_number,
                defaults={
                    "first_name": first_name,
                    "last_name": last_name or "—",
                    "email": None,
                    "notification_preference": "sms",
                    "is_active": True,
                    "notes": (
                        f"Imported from Excel ({country})"
                        if country
                        else "Imported from Excel"
                    ),
                },
            )

            if created:
                created_count += 1
            else:
                updated = False
                if not person.first_name and first_name:
                    person.first_name = first_name
                    updated = True
                if not person.last_name and last_name:
                    person.last_name = last_name
                    updated = True
                if country and (
                    not person.notes or "Imported from Excel" in (person.notes or "")
                ):
                    person.notes = (person.notes or "") + f" | Imported country: {country}"
                    updated = True
                if updated:
                    person.save()
                    updated_count += 1
                else:
                    skipped_count += 1  # Already existed with data

        self.stdout.write(
            self.style.SUCCESS(
                f"Import complete. Created: {created_count}, "
                f"Updated: {updated_count}, Skipped: {skipped_count}"
            )
        )

    def _split_name(self, name: str):
        """Split full name into first and last name."""
        parts = name.split()
        if not parts:
            return "", ""
        first_name = parts[0]
        last_name = " ".join(parts[1:]) if len(parts) > 1 else ""
        return first_name, last_name

    def _to_e164(self, digits: str, country: str | None) -> str:
        """
        Convert raw digits + country to an E.164 number string (+countrycode...).

        We handle Ghana and Togo specially, and fall back to Ghana if unsure.
        """

        country = (country or "").strip().lower()

        # Ghana numbers – docs say 233XXXXXXXXX
        if "ghana" in country:
            core = digits[-9:]  # last 9 digits
            return f"+233{core}"

        # Togo numbers – use +228 (based on country column)
        if "togo" in country:
            core = digits[-8:]  # typical Togo mobile length
            return f"+228{core}"

        # Fallback: if digits already start with a country code (233..., 228..., etc.)
        if digits.startswith("233") or digits.startswith("228"):
            return f"+{digits}"

        # As a last resort, treat as Ghana local number
        core = digits[-9:]
        return f"+233{core}"


