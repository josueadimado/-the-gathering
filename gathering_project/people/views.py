from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import JsonResponse
import re
import openpyxl
from .models import Person
from .forms import PersonRegistrationForm, PersonAdminForm, ExcelImportForm

# Create your views here.

def register(request):
    """Public registration form - no login required."""
    if request.method == 'POST':
        form = PersonRegistrationForm(request.POST)
        if form.is_valid():
            person = form.save()
            messages.success(request, f'Thank you {person.first_name}! You have been registered successfully.')
            return redirect('people:register_success')
    else:
        form = PersonRegistrationForm()
    
    return render(request, 'people/register.html', {'form': form})


def register_success(request):
    """Success page after registration."""
    return render(request, 'people/register_success.html')


@login_required
def person_list(request):
    """List all registered people (admin only)."""
    search_query = request.GET.get('search', '')
    people = Person.objects.all()
    
    if search_query:
        people = people.filter(
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(phone_number__icontains=search_query) |
            Q(email__icontains=search_query)
        )

    # Pagination: 25 people per page
    paginator = Paginator(people, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'people': page_obj,  # for backward compatibility in template
        'page_obj': page_obj,
        'search_query': search_query,
    }
    return render(request, 'people/person_list.html', context)


@login_required
def person_detail(request, pk):
    """View details of a specific person (admin only)."""
    person = get_object_or_404(Person, pk=pk)
    return render(request, 'people/person_detail.html', {'person': person})


@login_required
def person_update(request, pk):
    """Update person details (admin only)."""
    person = get_object_or_404(Person, pk=pk)
    
    if request.method == 'POST':
        form = PersonAdminForm(request.POST, instance=person)
        if form.is_valid():
            person = form.save()
            messages.success(request, f'{person.get_full_name()}\'s details have been updated successfully!')
            return redirect('people:detail', pk=person.pk)
    else:
        form = PersonAdminForm(instance=person)
    
    return render(request, 'people/person_update.html', {'form': form, 'person': person})


@login_required
def admin_register(request):
    """Admin-only registration form - simpler interface."""
    if request.method == 'POST':
        form = PersonAdminForm(request.POST)
        if form.is_valid():
            person = form.save()
            messages.success(request, f'{person.get_full_name()} has been registered successfully!')
            return redirect('people:list')
    else:
        form = PersonAdminForm()
    
    return render(request, 'people/admin_register.html', {'form': form})


@login_required
def import_excel(request):
    """Upload and import people from Excel file."""
    if request.method == 'POST':
        form = ExcelImportForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']
            
            # Check file extension
            if not excel_file.name.endswith(('.xlsx', '.xls')):
                messages.error(request, 'Please upload a valid Excel file (.xlsx or .xls)')
                return render(request, 'people/import_excel.html', {'form': form})
            
            try:
                # Process the Excel file
                result = process_excel_file(excel_file)
                
                if result['success']:
                    messages.success(
                        request,
                        f"Import complete! Created: {result['created']}, "
                        f"Updated: {result['updated']}, Skipped: {result['skipped']}"
                    )
                    return redirect('people:list')
                else:
                    messages.error(request, f"Error importing file: {result['error']}")
            except Exception as e:
                messages.error(request, f"An error occurred: {str(e)}")
    else:
        form = ExcelImportForm()
    
    return render(request, 'people/import_excel.html', {'form': form})


def process_excel_file(excel_file):
    """Process uploaded Excel file and import people."""
    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        ws = wb.active
        
        # Read headers from first row
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = [str(h).strip() if h is not None else "" for h in header_row]
        header_map = {h.lower(): idx for idx, h in enumerate(headers)}
        
        # Find column indices
        def get_index(*candidates):
            for cand in candidates:
                idx = header_map.get(cand.lower())
                if idx is not None:
                    return idx
            return None
        
        name_idx = get_index("name")
        country_idx = get_index("country", "country ")
        contact_idx = get_index("contact")
        
        if name_idx is None or contact_idx is None:
            return {
                'success': False,
                'error': f"Expected columns 'Name' and 'Contact' in the header row. Found: {headers}"
            }
        
        created_count = 0
        updated_count = 0
        skipped_count = 0
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            name = (row[name_idx] or "").strip() if row[name_idx] else ""
            country = (
                (row[country_idx] or "").strip()
                if country_idx is not None and row[country_idx]
                else ""
            )
            contact_raw = row[contact_idx]
            
            if not name and not contact_raw:
                continue
            
            if not contact_raw:
                skipped_count += 1
                continue
            
            # Normalize contact to digits only
            contact_str = re.sub(r"\D", "", str(contact_raw))
            if not contact_str:
                skipped_count += 1
                continue
            
            phone_number = to_e164(contact_str, country)
            first_name, last_name = split_name(name)
            
            person, created = Person.objects.get_or_create(
                phone_number=phone_number,
                defaults={
                    "first_name": first_name,
                    "last_name": last_name or "—",
                    "email": None,
                    "notification_preference": "sms",
                    "is_active": True,
                    "notes": (
                        f"Imported from Excel ({country})"
                        if country
                        else "Imported from Excel"
                    ),
                },
            )
            
            if created:
                created_count += 1
            else:
                updated = False
                if not person.first_name and first_name:
                    person.first_name = first_name
                    updated = True
                if not person.last_name and last_name:
                    person.last_name = last_name
                    updated = True
                if country and (
                    not person.notes or "Imported from Excel" in (person.notes or "")
                ):
                    person.notes = (person.notes or "") + f" | Imported country: {country}"
                    updated = True
                if updated:
                    person.save()
                    updated_count += 1
                else:
                    skipped_count += 1
        
        return {
            'success': True,
            'created': created_count,
            'updated': updated_count,
            'skipped': skipped_count
        }
    
    except Exception as e:
        return {
            'success': False,
            'error': str(e)
        }


def split_name(name: str):
    """Split full name into first and last name."""
    parts = name.split()
    if not parts:
        return "", ""
    first_name = parts[0]
    last_name = " ".join(parts[1:]) if len(parts) > 1 else ""
    return first_name, last_name


def to_e164(digits: str, country: str | None) -> str:
    """Convert raw digits + country to an E.164 number string."""
    country = (country or "").strip().lower()
    
    # Ghana numbers
    if "ghana" in country:
        core = digits[-9:]
        return f"+233{core}"
    
    # Togo numbers
    if "togo" in country:
        core = digits[-8:]
        return f"+228{core}"
    
    # If digits already start with a country code
    if digits.startswith("233") or digits.startswith("228"):
        return f"+{digits}"
    
    # Fallback: treat as Ghana local number
    core = digits[-9:]
    return f"+233{core}"

